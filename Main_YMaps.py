import os
import re
import sys
import random
import asyncio
import openpyxl
from typing import List
from openpyxl import Workbook
from deep_translator import GoogleTranslator
from playwright.async_api import async_playwright
from playwright.async_api import TimeoutError as PlaywrightTimeoutError


class YMapsParse:
    def __init__(self, keyword: str, city, max_num_firm: int, gui_url_work: bool):
        self.keyword = keyword  # Ищем по ключевому слову
        self.city = city  # Ищем в определённом городе
        self.max_num_firm = max_num_firm  # Максимальное количество фирм
        self.data_saving = "ymaps_parse_results/ymaps_data.xlsx"
        self.list_of_companies = []  # сюда добавляем списки из __get_firm_data, чтобы потом ввести их в xlsx
        self.start_row = 2
        self.gui_url_work = gui_url_work
        self.page2 = None
        self.page = None
        self.context = None
        self.stop_requested = False

        self.warning_message()

    async def random_delay(self, min_seconds=1, max_seconds=3):
        """Случайная задержка между действиями"""
        await asyncio.sleep(random.uniform(min_seconds, max_seconds))

    async def translate_text(self, city):
        """Переводим город на английский для удобства"""
        # Проверяем, является ли слово английским (только латинские буквы)
        right_city = re.sub(r"[^a-zA-Z\s]", "", city).strip()
        is_english = bool(re.match(r"^[a-zA-Z\s\-]+$", right_city))

        """try:
            return city_mapping[city]
        except:
            pass"""
        right_city = re.sub(r"[^а-яА-Я\s]", "", city).strip()
        if is_english:
            # Если уже английское слово, просто форматируем
            city_clean = "-".join(right_city.split())
            return city_clean.lower()
        else:
            # Если русское слово - переводим
            self.translator = GoogleTranslator(source="ru", target="en")
            a = await asyncio.to_thread(self.translator.translate, city)
            a = "-".join(a.split()) 
            return a.lower()

    async def __get_links(self, update_callback=None) -> List[str]:
        """Извлекаем ссылки на организации, находящиеся на странице"""
        print("Собираем ссылки на организации с текущей страницы...")
        self.list_of_companies = []

        try:
            container = await self.page.wait_for_selector(".scroll__container", timeout=5000)

            previous_count = 0
            max_scrolls = 50

            for _ in range(max_scrolls):

                # Получаем текущее количество карточек
                elements = await self.page.query_selector_all(".link-wrapper")
                current_count = len(elements)

                print(f"Найдено карточек: {current_count}")

                # Если достигли лимита
                if current_count >= self.max_num_firm:
                    print("Достигнут лимит фирм")
                    break

                # Если новые карточки не загружаются — прекращаем
                if current_count == previous_count:
                    print("Новые карточки больше не загружаются")
                    break

                previous_count = current_count

                # Скроллим вниз
                await container.evaluate("el => el.scrollBy(0, el.scrollHeight)")
                await self.page.wait_for_timeout(1500)

            # После завершения скролла получаем финальный список ссылок
            links = await self.page.eval_on_selector_all(
                ".link-wrapper",
                "elements => elements.map(el => el.href)"
            )

            for link in links[:self.max_num_firm]:
                if self.stop_requested:
                    break
                link = link.rstrip("/gallery/")
                firm_data = await self.__get_firm_data(url=link)

                if self.phone_text != "---" or (
                    self.phone_text == "---" and self.site_text != "Нет ссылки на сайт"
                ):
                    self.list_of_companies.append(firm_data)
                if update_callback:
                        update_callback(firm_data[1:-1])

        except PlaywrightTimeoutError:
            print("Контейнер не найден")

    async def __get_firm_data(self, url: str):
        """Берем данные фирмы: название, телефон, сайт"""
        if self.stop_requested:
            return []
        self.page2 = await self.context.new_page()  # Создаем новую страницу
        try:
            await self.page2.goto(url=url)  # Переходим на неё
            if self.stop_requested:
                return []
            await self.random_delay(0.5, 1)

            # ИНИЦИАЛИЗИРУЕМ ВСЕ ПЕРЕМЕННЫЕ ЗНАЧЕНИЯМИ ПО УМОЛЧАНИЮ ДО ПОИСКА
            name_firm_text = "Название не найдено"
            address_text = "Адрес не указан"
            phone_text = "Телефон не найден"
            site_text = "Нет ссылки на сайт"

            # Название компании
            try:
                name_firm_element = await self.page2.query_selector(".orgpage-header-view__header")
                if name_firm_element:
                    name_firm_text = await name_firm_element.text_content()
                    name_firm_text = name_firm_text.strip()
            except Exception as e:
                print(f"Ошибка при получении названия: {e}")

            # Адрес
            try:
                address_element = await self.page2.query_selector(".business-contacts-view__address-link")
                if address_element:
                    address_text = await address_element.text_content()
                    address_text = address_text.strip()
            except Exception as e:
                print(f"Ошибка при получении адреса: {e}")

            # Категории
            try:
                category_element = await self.page2.query_selector_all(".breadcrumbs-view__breadcrumb")
                if category_element:
                    category_text = await category_element[-1].text_content()
                    category_text = category_text.strip()
                else:
                    self.phone_text = "---"
            except Exception as e:
                print(f"Ошибка при получении категории: {e}")

            # Номер телефона
            try:
                phone_container = await self.page2.query_selector(".card-phones-view__number")
                if phone_container:
                    phone_text = await phone_container.text_content()
                    self.phone_text = phone_text.rstrip("Показать телефон")
                else:
                    self.phone_text = "---"
            except Exception as e:
                print(f"Ошибка при получении телефона: {e}")
                self.phone_text = "---"

            # Название сайта
            try:
                site_element = await self.page2.query_selector(".business-urls-view__text")
                if site_element:
                    site_text = await site_element.text_content()
                    self.site_text = site_text.strip()
                else:
                    self.site_text = "Нет ссылки на сайт"
            except Exception as e:
                print(f"Ошибка при получении сайта: {e}")
                site_text = "Нет ссылки на сайт"
            await self.random_delay(0.5, 1)

            # Возвращаем ВСЕ переменные (они теперь точно определены)
            return [
                url,
                name_firm_text,
                category_text,
                address_text,
                self.phone_text,
                self.site_text,
                "-",
            ]
        finally: 
            await self.page2.close()

    async def check_xlsx(self, update_callback):
        """Функция для создания заготовки под xlsx файл"""
        # Проверки, есть ли папка, если нет, то создаем
        try:
            if os.path.exists(self.data_saving):
                os.remove(self.data_saving)
        except Exception as e:
            print(f"Ошибка: {e}")
            if update_callback:
                update_callback("Ошибка: Проверьте/Закройте файл 'ymaps_data.xlsx'")
            sys.exit(1)
        os.makedirs("ymaps_parse_results", exist_ok=True)

        # Создать новый файл (старый удаляется при включении программы)
        self.wb = Workbook()
        self.ws = self.wb.active

        # Добавляем заголовки
        headers = ["URL", "Название", "Категория", "Адрес", "Телефон", "Сайт"]
        for col, header in enumerate(headers, start=1):
            self.ws.cell(row=1, column=col, value=header)

    async def data_output_to_xlsx(self, get_firm_data, update_callback):
        """Выводим данные в файл xlsx"""
        # Открыть существующий файл
        if os.path.exists(self.data_saving):
            self.wb = openpyxl.load_workbook(self.data_saving)
            self.ws = self.wb.active
        # Цикл по данным фирм
        for firm_data in get_firm_data:  # firm_data - это список ['URL', 'Название', 'Телефон', 'Сайт']
            # Запись каждой строки в Excel
            for col, value in enumerate(firm_data, start=1):
                self.ws.cell(row=self.start_row, column=col, value=value)
            self.start_row += 1  # Перейти на следующую строку

        # Сохранить файл
        self.wb.save(self.data_saving)
        firm_data_list = list(map(lambda x: x[1:-1], self.list_of_companies))
        print(f"Записано {len(get_firm_data)} строк в файл data.xlsx")
        
    async def stop(self):
        self.stop_requested = True

        try:
            if self.page and not self.page.is_closed():
                await self.page.close()
        except:
            pass

        try:
            if self.context:
                await self.context.close()
        except:
            pass

    async def get_random_user_agent(self):
        """Скрываем автоматизацию с помощью захода с разных систем"""
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
        ]
        return random.choice(user_agents)

    def warning_message(self):
        print("\n" + "=" * 50)
        print("EDUCATIONAL USE ONLY - NO WARRANTY PROVIDED")
        print("This parser may violate Terms of Service.")
        print("Use only for learning web scraping techniques.")
        print("Author not responsible for any legal consequences.")
        print("=" * 50 + "\n")

    async def parse_main(self, update_callback=None):
        """Парсинг сайта"""
        async with async_playwright() as playwright:
            try:
                await self.check_xlsx(update_callback)
                browser = await playwright.chromium.launch(headless=False)  # headless=True - без графического итерфейса
                self.context = await browser.new_context(
                    user_agent=await self.get_random_user_agent(),
                    locale="ru-RU",
                    timezone_id="Europe/Moscow",
                )  # По типу вкладок инкогнито
                self.page = await self.context.new_page()  # Новая страница, создается в контексте
                if self.gui_url_work:
                    await self.page.goto(
                        self.keyword,
                        wait_until="domcontentloaded",
                    )  # Переходим по адресу с переведенным городом
                else:
                    await self.page.goto(
                        f"https://yandex.ru/maps/1/a/search/{self.city}, {self.keyword}",
                        wait_until="domcontentloaded",
                    )  # Переходим по адресу с переведенным городом

                await self.random_delay(3, 4)  # Задержка для загрузки страницы

                # Собираем данные с задержками
                while self.ws.max_row < self.max_num_firm:
                    if self.stop_requested:
                        print("Парсинг остановлен")
                        break
                    if self.ws.max_row - 1 != 0:
                        print(f"Записанных фирм в xlsx: {self.ws.max_row - 1}")
                    await self.__get_links(update_callback)  # Ищем ссылки и данные организаций
                    await self.data_output_to_xlsx(self.list_of_companies, update_callback)  # Записываем данные в Excel
                    # Имитация просмотра страницы
                    await self.random_delay(1, 2)

                    # Переход на следующую страницу с проверкой
                    next_button = await self.page.query_selector('[style="transform: rotate(-90deg);"]')
                    if next_button:
                        color = await next_button.evaluate(
                            """el => window.getComputedStyle(el).color"""
                        )
                    if next_button and color == "rgb(0, 114, 206)":
                        await self.random_delay(1, 2)
                        await next_button.click()
                        await self.random_delay(3, 4.5)  # Ждем загрузки следующей страницы
                    else:
                        break  # Больше нет страниц
                else:
                    if self.page2:
                        await self.page2.close()
                    await browser.close()

                print(f"Записано {self.ws.max_row - 1} организаций")
            except Exception as e:
                error_msg = f"Произошла ошибка: {e}"
                print(error_msg)
                if update_callback:
                    update_callback(error_msg)
                raise


async def main():
    parser = YMapsParse(
        keyword="Велопрокат", city="Самара", max_num_firm=5, gui_url_work=False
    )
    await parser.parse_main()


if __name__ == "__main__":
    asyncio.run(main())
